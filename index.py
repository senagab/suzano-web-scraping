import tkinter as tk
from tkinter import messagebox
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
import time

def rodar_scraping():
    try:
        # Caminho do arquivo Excel
        caminho_planilha = r"C:\Users\senag\Documents\0 - Projects\Banco\banco.xlsx"

        # Carregar a planilha
        wb = load_workbook(caminho_planilha)
        sheet = wb.active

        # Inicializar o WebDriver
        service = Service(r"C:\Users\senag\Documents\0 - Projects\Freelance\Suzano Web Scraping\chromedriver.exe")
        driver = webdriver.Chrome(service=service)
        wait = WebDriverWait(driver, 10)  # Espera até 10 segundos para os elementos aparecerem

        # Abrir o site e maximizar a janela
        driver.get("https://suzano-web-scraping.vercel.app/")
        driver.maximize_window()  # Maximizar a janela

        time.sleep(3)  # Garantir carregamento inicial da página

        # Iterar sobre as linhas da planilha, pulando o cabeçalho
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Verifica se todos os campos da linha estão vazios (None ou string vazia)
            if all(cell is None or str(cell).strip() == "" for cell in row):
                print("Linha vazia encontrada. Encerrando o processo...")
                break  # Sai do loop quando uma linha vazia for encontrada

            # Garantir que não há valores None
            campo_status = str(row[0]) if row[0] else ""
            campo_chegada = str(row[1]) if row[1] else ""
            campo_saida = str(row[2]) if row[2] else ""
            campo_veiculo = str(row[3]) if row[3] else ""
            campo_modelo_veiculo = str(row[4]) if row[4] else ""
            campo_data = str(row[5]) if row[5] else ""  # Converter data para string
            campo_motorista = str(row[6]) if row[6] else ""

            # Preencher os campos do formulário com espera para garantir que os elementos existem
            wait.until(EC.presence_of_element_located((By.ID, "campo-status"))).send_keys(campo_status)
            wait.until(EC.presence_of_element_located((By.ID, "campo-chegada"))).send_keys(campo_chegada)
            wait.until(EC.presence_of_element_located((By.ID, "campo-saida"))).send_keys(campo_saida)
            wait.until(EC.presence_of_element_located((By.ID, "campo-veiculo"))).send_keys(campo_veiculo)
            wait.until(EC.presence_of_element_located((By.ID, "campo-modelo-veiculo"))).send_keys(campo_modelo_veiculo)
            wait.until(EC.presence_of_element_located((By.ID, "campo-data"))).send_keys(campo_data)
            wait.until(EC.presence_of_element_located((By.ID, "campo-motorista"))).send_keys(campo_motorista)

            # Pressionar Enter para enviar o formulário
            driver.find_element(By.ID, "campo-motorista").send_keys(Keys.RETURN)

            time.sleep(2)  # Pequeno delay entre envios para evitar bloqueios

        messagebox.showinfo("Sucesso", "Formulários preenchidos com sucesso!")
    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro: {e}")
    finally:
        driver.quit()


# Criar a interface gráfica
root = tk.Tk()
root.title("Automatização de Preenchimento de Formulários")
root.geometry("400x200")  # Tamanho da janela

# Criar o botão
botao_rodar = tk.Button(root, text="Iniciar", command=rodar_scraping, width=20, height=2)
botao_rodar.pack(pady=50)

# Rodar a interface gráfica
root.mainloop()
